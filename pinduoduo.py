# -*- coding: utf-8 -*-
import requests
import os
import random
import time
import re
from urllib.parse import quote
from lxml import etree
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from proxies import Proxies


# 电脑版拼多多商城：https://youhui.pinduoduo.com
# 手机版拼多多商城：http://www.mobile.yangkeduo.com

def ua():
    ua_list = [
         'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.1 (KHTML, like Gecko) Chrome/14.0.835.163 Safari/535.1',
         'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:6.0) Gecko/20100101 Firefox/6.0',
         'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',
         'Opera/9.80 (Windows NT 6.1; U; zh-cn) Presto/2.9.168 Version/11.50',
         'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 2.0.50727; SLCC2; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.3; .NET4.0C; Tablet PC 2.0; .NET4.0E)',
         'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1)',
         'Opera/9.80 (X11; Linux i686; U; ru) Presto/2.8.131 Version/11.11',
         "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.67 Safari/537.36",
         "Mozilla/5.0 (X11; OpenBSD i386) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.125 Safari/537.36",
         "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1944.0 Safari/537.36",
         "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.3319.102 Safari/537.36",
         "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.2309.372 Safari/537.36",
         "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.2117.157 Safari/537.36"]
    ua = random.choice(ua_list)
    return ua


# 获取店铺url
def shopid(goodsid):
    goodsid = re.findall('=(\d+)&', goodsid)[0]
    url = 'https://youhui.pinduoduo.com/network/api/goods/queryByGoodsId'
    data = {'goodsIds': [str(goodsid)]}
    headers = {'User-Agent': ua()}
    rep = requests.post(url, json=data, headers=headers)
    # 获取店铺id
    shop_id = rep.json()['result']['goodsDetails'][0]['mallId']
    shop_url = 'http://www.mobile.yangkeduo.com/mall_page.html?mall_id=' + str(shop_id)
    return shop_url


# 查找详细信息页面
def link(keyw, url, proxies):
    shop_url = shopid(url)
    key = quote(keyw)
    # url = 'http://www.mobile.yangkeduo.com/goods.html?goods_id=8441010&gallery_id=04a9394e5a690f63e38db3aa18be8f4f'
    heards = {
        'User-Agent': ua(),
        'Referer': 'http://www.mobile.yangkeduo.com/search_result.html?search_key='+key,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Host': 'www.mobile.yangkeduo.com',
        'Upgrade-Insecure-Requests': '1',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'en-US,en;q=0.9,zh-CN;q=0.8,zh;q=0.7',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive'
    }
    rep = requests.get(url, headers=heards, proxies=proxies)
    # rep = requests.get(url, headers=heards)
    # print(rep.text)
    html = etree.HTML(rep.text)
    # 店铺名称
    shop_name = ''.join(html.xpath('//*[@class="goods-mall-name"]/text()'))
    # 商品量和销售量
    num = html.xpath('//*[@class="from-server"]/text()')
    goods_num = int(num[0][6:])
    sale_num = num[1][4:-1]
    print(shop_name,goods_num,sale_num)
    # 判断是否满足条件，满足的写入文件。(抓取店铺商品超过200同时销量超过一百的抓取字段为：关键词，店铺名称，店铺商品数量，店铺总销量)
    if goods_num >= 200:
        if '万' in sale_num:
            wb = load_workbook(keyw+'.xlsx')
            ws = wb.active
            ws.append([keyw, shop_name, goods_num, sale_num, shop_url])
            wb.save(keyw + '.xlsx')
    # print(shop_name,goods_num,sale_num)
        if '万' not in sale_num:
            sale_num = int(sale_num)
            if sale_num >= 100:
                wb = load_workbook(key + '.xlsx')
                ws = wb.active
                ws.append([key, shop_name, goods_num, sale_num, shop_url])
                wb.save(key + '.xlsx')
    #         print(shop_name,goods_num,sale_num)


# 关键词搜索函数
def search(key, sort_type, page_num, proxies):
    url_list = []
    url = 'https://jinbao.pinduoduo.com/network/api/common/goodsList'
    word = quote(key)
    headers = {
        'referer': 'https://youhui.pinduoduo.com/search/landing?keyword='+word,
        'origin': 'https://youhui.pinduoduo.com',
        'User-Agent': ua(),
        # 'cookie': 'api_uid=rBQQsFvikC5qpV2/G0+jAg==; _ga=GA1.2.1411405322.1541574704; _gid=GA1.2.629367774.1542348220; pt_691035c2=uid=UtBk/sZrPFQoVvk2e3PaCA&nid=0&vid=9hqZm9ivU4gxYo8Rwa-JvQ&vn=10&pvn=1&sact=1542353135482&to_flag=0&pl=cXwVzCBtg4JBGEEuanQrLQ*pt*1542353135482; pt_s_691035c2=vt=1542353135482&cad=',
        'accept': 'application/json, text/plain, */*',
        'accept-encoding': 'gzip, deflate, br',
        'content-length': '77',
        'content-type': 'application/json; charset=UTF-8'
    }
    data = {"keyword": key, "sortType": sort_type, "withCoupon": 0, "pageNumber": page_num, "pageSize": 60}

    rep = requests.post(url, json=data, headers=headers, proxies=proxies)
    # rep = requests.post(url, json=data, headers=headers)
    print(rep.json())
    data = rep.json()
    # print(len(data["result"]["goodsList"]))
    for i in range(len(data["result"]["goodsList"])):
        goodsid = data["result"]["goodsList"][i]["goodsId"]
        goodsthumbnailurl = data["result"]["goodsList"][i]['goodsThumbnailUrl'][-37:-5]
        url = 'http://www.mobile.yangkeduo.com/goods.html?goods_id='+str(goodsid)+'&gallery_id='+goodsthumbnailurl
        url_list.append(url)
    # print(len(url_list),url_list)
    return url_list


def reatefile(key):
    if os.path.exists(key+".xlsx"):
        pass
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['关键词', '店铺名称', '店铺商品数量', '店铺总销量', '店铺连接'])
        wb.save(key+".xlsx")


def start(key, sort_type):
    reatefile(key)
    proxies = Proxies().main()
    # proxies={}
    for page_num in range(1, 15):
        try:
            url_list = search(key, sort_type, page_num, proxies)
            # print('****',url_list)
            for url in url_list:
                link(key, url, proxies)
                time.sleep(random.uniform(1, 3))
        except Exception as e:
            print('* error *',e)
            if len(url_list) == 0:
                proxies = Proxies().main()




if __name__ == '__main__':
    print('排序方式分为：')
    print('默认排序输入:0')
    print('价格从低到高:3')
    print('价格从高到低:4')
    print('销量从低到高:5')
    print('销量从高到低:6')

    key = input('关键词：')
    try:
        sort_type = int(input('排序方式为：'))
        if sort_type == 1:
            start(key, sort_type)
        elif sort_type == 3:
            start(key, sort_type)
        elif sort_type == 4:
            start(key, sort_type)
        elif sort_type == 5:
            start(key, sort_type)
        elif sort_type == 6:
            start(key, sort_type)
    except Exception:
        pass
        # print('排序方式输入错误')
